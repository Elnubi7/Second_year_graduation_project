# excel_utils.py
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

class ExcelLogger:
    CLASSES = ['Eating', 'Working', 'Sleeping', 'Speaking on phone', 'Working']

    def __init__(self, path: str):
        self.path = path
        if os.path.exists(path):
            self.wb = load_workbook(path)
        else:
            self.wb = Workbook()
        self.ws = self._get_today_sheet()

    def _get_today_sheet(self):
        today = datetime.now().strftime('%Y-%m-%d')
        if today not in self.wb.sheetnames:
            ws = self.wb.create_sheet(today)
            self._setup_sheet(ws)
        else:
            ws = self.wb[today]
        return ws

    def _setup_sheet(self, ws):
        headers = ['Desk'] + self.CLASSES
        ws.append(headers)
        bold_center = Font(bold=True)
        center = Alignment(horizontal='center')
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=col)
            cell.font = bold_center
            cell.alignment = center
        for i, h in enumerate(headers, start=1):
            ws.column_dimensions[chr(64+i)].width = max(len(h)+2, 12)

    def append_row(self, row: list):
        self.ws.append(row)

    def finalize(self):
        color_map = {
            'Eating': 'FFEB9C',
            'Working': 'C6EFCE',
            'Sleeping': 'D9D9D9',
            'Speaking on phone': 'BDD7EE'
        }
        for row in self.ws.iter_rows(min_row=2, max_col=len(self.CLASSES)+1):
            max_val, max_idx = -1, None
            for idx, cell in enumerate(row[1:], start=1):
                try:
                    v = int(cell.value or 0)
                except:
                    v = 0
                if v > max_val:
                    max_val, max_idx = v, idx
            if max_idx:
                fill = PatternFill(
                    start_color=color_map[self.CLASSES[max_idx-1]],
                    end_color=color_map[self.CLASSES[max_idx-1]],
                    fill_type='solid'
                )
                row[max_idx].fill = fill

    def save(self):
        self.wb.save(self.path)